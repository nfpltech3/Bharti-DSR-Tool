# ─────────────────────────────────────────────────────────────────────────────
# BRANDING CONSTANTS (NAGARKOT STANDARD)
# ─────────────────────────────────────────────────────────────────────────────
BRAND_BLUE   = "#1F3F6E"
BRAND_RED    = "#D8232A"
TEXT_DARK    = "#1E1E1E"
MUTED_GRAY   = "#6B7280"
LIGHT_BG     = "#F4F6F8"
WHITE        = "#FFFFFF"
BORDER_GRAY  = "#E5E7EB"
HOVER_BLUE   = "#2A528F"

HIGHLIGHT_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC",
                              fill_type="solid")
HEADER_FILL    = PatternFill(start_color="1F3F6E", end_color="1F3F6E",
                              fill_type="solid")
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT    = Font(size=9)
BOLD_FONT      = Font(bold=True, size=9)

# Columns to skip during comparison (identifiers / row counters)
SKIP_COLUMNS = {"s.n", "s.n.", "invoice no", "invoice no.", "sn", "sr no",
                "sr. no", "sr.no"}

# ─────────────────────────────────────────────────────────────────────────────
# UTILITY FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def norm_col(name: str) -> str:
    """Normalize column name: collapse whitespace + strip + lowercase."""
    return re.sub(r'\s+', ' ', str(name)).strip().lower()


def make_invoice_key(raw) -> str:
    """
    Extract a canonical invoice key from a (possibly multi-line) cell value.
    """
    if pd.isna(raw):
        return ""
    parts = [p.strip() for p in str(raw).split('\n') if p.strip()]
    return " | ".join(sorted(set(parts)))


DATE_FORMATS = [
    "%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%d-%m-%Y",
    "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f",
    "%m/%d/%Y", "%d %b %Y", "%d %B %Y",
    "%d-%b-%y", "%d/%m/%y",
]

def _normalize_single(s: str) -> str:
    """Normalize a single (non-multiline) string value."""
    s = s.strip()
    if not s or s in ('nan', 'NaT', 'None'):
        return ""
    # Try parsing as date
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).strftime("%d-%b-%Y").upper()
        except ValueError:
            pass
    # Numeric: remove trailing .0
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return s
    except ValueError:
        pass
    return re.sub(r'\s+', ' ', s).upper()


def normalize_value(val) -> str:
    """
    Convert any value to a comparable string.
    """
    if pd.isna(val) or val is None or str(val).strip() in ('', 'nan', 'NaT'):
        return ""

    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%d-%b-%Y").upper()

    s = str(val).strip()
    if not s:
        return ""

    if '\n' in s:
        lines = [l.strip() for l in s.split('\n') if l.strip()]
        normed = list(dict.fromkeys(_normalize_single(l) for l in lines
                                    if _normalize_single(l)))
        if not normed:
            return ""
        return normed[0] if len(normed) == 1 else " | ".join(normed)

    return _normalize_single(s)


def map_sheets(gen_sheets: list, man_sheets: list) -> dict:
    """
    Return {gen_sheet_name: man_sheet_name} for sheets that match.
    """
    man_norm = {s.strip().lower(): s for s in man_sheets}
    result = {}
    for gs in gen_sheets:
        key = gs.strip().lower()
        if key in man_norm:
            result[gs] = man_norm[key]
    return result


def find_invoice_col(df: pd.DataFrame) -> str | None:
    """Find the invoice number column regardless of exact name."""
    for col in df.columns:
        n = norm_col(col)
        if n in ("invoice no", "invoice no."):
            return col
    return None


# ─────────────────────────────────────────────────────────────────────────────
# AUDIT ENGINE
# ─────────────────────────────────────────────────────────────────────────────
def run_audit(gen_path: str, man_path: str, progress_cb=None) -> dict:
    """
    Core audit logic.
    """
    def progress(msg):
        if progress_cb:
            progress_cb(msg)

    progress("Reading files...")
    gen_xl = pd.ExcelFile(gen_path)
    man_xl = pd.ExcelFile(man_path)

    sheet_map = map_sheets(gen_xl.sheet_names, man_xl.sheet_names)
    if not sheet_map:
        raise ValueError(
            "No matching sheets found between the two files.\n"
            f"Generated sheets: {gen_xl.sheet_names}\n"
            f"Manual sheets: {man_xl.sheet_names}")

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir     = os.path.dirname(gen_path)
    output_path = os.path.join(out_dir, f"DSR_Audit_Report_{timestamp}.xlsx")

    total_checked  = 0
    total_mismatch = 0
    total_unmatched = 0
    detail_rows    = []
    sheets_audited = 0

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    ws_summary = wb_out.create_sheet("AUDIT SUMMARY", 0)

    for gen_sheet, man_sheet in sheet_map.items():
        progress(f"Auditing sheet: {gen_sheet}...")

        gen_df = pd.read_excel(gen_path, sheet_name=gen_sheet,
                               dtype=str, keep_default_na=False)
        man_df = pd.read_excel(man_path, sheet_name=man_sheet,
                               dtype=str, keep_default_na=False)

        gen_inv_col = find_invoice_col(gen_df)
        man_inv_col = find_invoice_col(man_df)
        if not gen_inv_col or not man_inv_col:
            progress(f"  Skipping '{gen_sheet}': no Invoice No column found.")
            continue

        man_lookup = {}
        for _, row in man_df.iterrows():
            key = make_invoice_key(row[man_inv_col])
            if key:
                man_lookup[key] = {norm_col(c): row[c] for c in man_df.columns}

        man_col_norm_map = {norm_col(c): c for c in man_df.columns}
        col_xref = {}
        for gc in gen_df.columns:
            nk = norm_col(gc)
            if nk in man_col_norm_map and nk not in SKIP_COLUMNS:
                col_xref[gc] = nk

        ws = wb_out.create_sheet(gen_sheet[:31])
        sheets_audited += 1

        headers = list(gen_df.columns)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        sheet_checked  = 0
        sheet_mismatch = 0

        for ri, (_, gen_row) in enumerate(gen_df.iterrows(), 2):
            inv_key = make_invoice_key(gen_row[gen_inv_col])
            man_row = man_lookup.get(inv_key)

            for ci, gc in enumerate(headers, 1):
                raw_gen = gen_row[gc]
                cell    = ws.cell(row=ri, column=ci, value=raw_gen)
                cell.font      = NORMAL_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                if man_row is None or gc not in col_xref:
                    continue

                nk      = col_xref[gc]
                raw_man = man_row.get(nk, "")

                gen_norm     = normalize_value(raw_gen)
                man_norm_val = normalize_value(raw_man)

                if gen_norm == "" and man_norm_val == "":
                    continue

                if gen_norm != man_norm_val:
                    cell.fill = HIGHLIGHT_FILL
                    comment_text = f"Manual DSR value:\n{raw_man}"
                    comment = Comment(comment_text, "DSR Auditor")
                    comment.width  = 220
                    comment.height = 60
                    cell.comment   = comment
                    sheet_mismatch += 1
                    detail_rows.append(
                        (gen_sheet, inv_key, gc,
                         str(raw_gen), str(raw_man)))

            if inv_key:
                if man_row is not None:
                    sheet_checked += 1
                else:
                    total_unmatched += 1

        for ci, col_cells in enumerate(ws.columns, 1):
            max_len = max(
                (len(str(c.value).split('\n')[0]) if c.value else 0)
                for c in col_cells)
            ws.column_dimensions[get_column_letter(ci)].width = min(
                max(8, max_len + 2), 35)

        total_checked  += sheet_checked
        total_mismatch += sheet_mismatch
        progress(f"  {gen_sheet}: {sheet_checked} matched, "
                 f"{sheet_mismatch} mismatches")

    ws_detail = wb_out.create_sheet("MISMATCH DETAIL")
    detail_headers = ["Sheet", "Invoice Key", "Column", "Generated Value",
                      "Manual (Correct) Value"]
    for ci, h in enumerate(detail_headers, 1):
        c = ws_detail.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for ri, row in enumerate(detail_rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws_detail.cell(row=ri, column=ci, value=val)
            cell.font = NORMAL_FONT
            if ci == 4:
                cell.fill = HIGHLIGHT_FILL
    for ci in range(1, 6):
        ws_detail.column_dimensions[get_column_letter(ci)].width = [
            18, 35, 30, 30, 30][ci - 1]
    ws_detail.freeze_panes = "A2"

    _write_summary(ws_summary, total_checked, total_mismatch,
                   total_unmatched, sheets_audited, sheet_map,
                   detail_rows)

    progress("Saving report...")
    wb_out.save(output_path)

    return {
        "output_path":      output_path,
        "invoices_checked": total_checked,
        "mismatches":       total_mismatch,
        "sheets_audited":   sheets_audited,
        "unmatched_gen":    total_unmatched,
        "details":          detail_rows,
    }

def _write_summary(ws, checked, mismatches, unmatched,
                   sheets_audited, sheet_map, detail_rows):
    from openpyxl.utils import get_column_letter

    BLUE_FILL  = PatternFill("solid", fgColor="1F3F6E")
    GRAY_FILL  = PatternFill("solid", fgColor="F0F0F0")
    GREEN_FILL = PatternFill("solid", fgColor="D4EDDA")
    AMBER_FILL = PatternFill("solid", fgColor="FFF3CD")

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = "NAGARKOT DSR AUDIT REPORT"
    t.fill      = BLUE_FILL
    t.font      = Font(bold=True, color="FFFFFF", size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:E2")
    ts = ws["A2"]
    ts.value     = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ts.font      = Font(italic=True, color="888888", size=9)
    ts.alignment = Alignment(horizontal="center")

    kpis = [
        ("Sheets Audited",    sheets_audited, GRAY_FILL),
        ("Invoices Checked",  checked,        GRAY_FILL),
        ("Mismatches Found",  mismatches,
         AMBER_FILL if mismatches > 0 else GREEN_FILL),
        ("Not in Manual",     unmatched,      GRAY_FILL),
        ("Match Rate",
         f"{round((checked - mismatches) / max(checked, 1) * 100, 1)}%",
         GREEN_FILL if mismatches == 0 else AMBER_FILL),
    ]
    for ci, (label, val, fill) in enumerate(kpis, 1):
        lc = ws.cell(row=4, column=ci, value=label)
        lc.fill      = fill
        lc.font      = Font(bold=True, size=9)
        lc.alignment = Alignment(horizontal="center", vertical="bottom")
        vc = ws.cell(row=5, column=ci, value=val)
        vc.fill      = fill
        vc.font      = Font(bold=True, size=20)
        vc.alignment = Alignment(horizontal="center", vertical="top")

    # Mapping table
    ws.cell(row=7, column=1, value="SHEET MAPPING").font = BOLD_FONT
    for ri, (gs, ms) in enumerate(sheet_map.items(), 9):
        ws.cell(row=ri, column=1, value=gs)
        ws.cell(row=ri, column=2, value=ms)
        ws.cell(row=ri, column=3, value="Matched ✔")

    for ci in range(1, 6):
        ws.column_dimensions[get_column_letter(ci)].width = 28


# ─────────────────────────────────────────────────────────────────────────────
# GUI (NAGARKOT BRAND STANDARD)
# ─────────────────────────────────────────────────────────────────────────────
class DSRAuditorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Nagarkot DSR Auditor")
        self.root.state("zoomed")
        self.root.configure(bg=LIGHT_BG)

        self.gen_path = tk.StringVar()
        self.man_path = tk.StringVar()

        self._build_gui()

    def _build_gui(self):
        # ── HEADER ────────────────────────────────────────────────────────
        self.hdr = tk.Frame(self.root, bg=WHITE, height=70, 
                            highlightthickness=1, highlightbackground=BORDER_GRAY)
        self.hdr.pack(fill=tk.X, side=tk.TOP)
        self.hdr.pack_propagate(False)

        # Logo (Left, Height 20)
        try:
            from PIL import Image, ImageTk
            logo_path = os.path.join(os.getcwd(), "logo.png")
            if os.path.exists(logo_path):
                img = Image.open(logo_path)
                aspect = img.width / img.height
                img = img.resize((int(20 * aspect), 20), Image.Resampling.LANCZOS)
                self.logo_img = ImageTk.PhotoImage(img)
                lbl_logo = tk.Label(self.hdr, image=self.logo_img, bg=WHITE)
                lbl_logo.place(x=30, y=25)
        except Exception:
            pass

        # Title (Center Absolute)
        lbl_title = tk.Label(self.hdr, text="NAGARKOT DSR AUDITOR",
                             font=("Arial", 18, "bold"),
                             fg=BRAND_BLUE, bg=WHITE)
        lbl_title.place(relx=0.5, rely=0.45, anchor=tk.CENTER)
        
        lbl_subtitle = tk.Label(self.hdr, text="Data Integrity & Comparison Engine",
                                font=("Arial", 9), fg=MUTED_GRAY, bg=WHITE)
        lbl_subtitle.place(relx=0.5, rely=0.75, anchor=tk.CENTER)

        # ── BODY ──────────────────────────────────────────────────────────
        body = tk.Frame(self.root, bg=LIGHT_BG, padx=80, pady=40)
        body.pack(fill=tk.BOTH, expand=True)

        card = tk.Frame(body, bg=WHITE, highlightthickness=1, 
                        highlightbackground=BORDER_GRAY, padx=40, pady=40)
        card.pack(fill=tk.X)

        # File Inputs
        self._file_field(card, row=0, label="1. SELECT GENERATED DSR (SOURCE)",
                         var=self.gen_path, hint="Exported file from Parser App")
        self._file_field(card, row=1, label="2. SELECT MANUAL DSR (REFERENCE)",
                         var=self.man_path, hint="Team-maintained master Excel")

        # Skip columns
        tk.Label(card, text="SKIP COLUMNS (S.N, Status, etc.):", 
                 font=("Arial", 9, "bold"), bg=WHITE).pack(anchor="w", pady=(20, 5))
        self.skip_entry = ttk.Entry(card, width=60)
        self.skip_entry.insert(0, "S.N, S.N., Curent status, Remarks, Remarks.1")
        self.skip_entry.pack(anchor="w")

        # Action
        self.btn_run = tk.Button(card, text="   RUN AUDIT   ",
                                 font=("Arial", 12, "bold"),
                                 bg=BRAND_BLUE, fg=WHITE,
                                 activebackground=HOVER_BLUE,
                                 activeforeground=WHITE,
                                 relief=tk.FLAT, pady=12, padx=40,
                                 command=self._run)
        self.btn_run.pack(pady=(40, 10))

        self.lbl_progress = tk.Label(card, text="Ready to audit files", 
                                     font=("Arial", 9), fg=MUTED_GRAY, bg=WHITE)
        self.lbl_progress.pack()

        # Results
        self.res_frame = tk.Frame(body, bg=WHITE, highlightthickness=1, 
                                  highlightbackground=BORDER_GRAY, pady=20)
        self.res_frame.pack(fill=tk.X, pady=20)
        
        self.lbl_res = tk.Label(self.res_frame, text="Audit summary will appear here.",
                                font=("Arial", 10), bg=WHITE, fg=MUTED_GRAY)
        self.lbl_res.pack()

        self.btn_open = tk.Button(self.res_frame, text="OPEN AUDIT REPORT",
                                  font=("Arial", 9, "bold"), bg=WHITE, fg=BRAND_BLUE,
                                  relief=tk.GROOVE, padx=20, pady=5,
                                  command=self._open_report, state=tk.DISABLED)
        self.btn_open.pack(pady=(10, 0))

        # ── FOOTER ────────────────────────────────────────────────────────
        ftr = tk.Frame(self.root, bg=LIGHT_BG, height=40)
        ftr.pack(fill=tk.X, side=tk.BOTTOM)
        tk.Label(ftr, text="Nagarkot Forwarders Pvt Ltd. \u00A9",
                 font=("Arial", 8), fg=MUTED_GRAY, bg=LIGHT_BG).pack(side=tk.LEFT, padx=40, pady=10)

    def _file_field(self, parent, row, label, var, hint):
        f = tk.Frame(parent, bg=WHITE)
        f.pack(fill=tk.X, pady=10)
        tk.Label(f, text=label, font=("Arial", 9, "bold"), bg=WHITE).pack(anchor="w")
        tk.Label(f, text=hint, font=("Arial", 7), fg=MUTED_GRAY, bg=WHITE).pack(anchor="w", pady=(0, 5))
        
        row_f = tk.Frame(f, bg=WHITE)
        row_f.pack(fill=tk.X)
        tk.Entry(row_f, textvariable=var, state="readonly", width=60).pack(side=tk.LEFT, padx=(0, 10))
        tk.Button(row_f, text="Browse...", bg=BRAND_RED, fg=WHITE, 
                  relief=tk.FLAT, padx=15, command=lambda: self._pick(var)).pack(side=tk.LEFT)

    def _pick(self, var):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path: var.set(path)

    def _run(self):
        gen = self.gen_path.get()
        man = self.man_path.get()
        if not gen or not man:
            messagebox.showwarning("Incomplete", "Please select both files.")
            return

        cols = [norm_col(c) for c in self.skip_entry.get().split(",") if c.strip()]
        SKIP_COLUMNS.update(cols)

        self.btn_run.config(text="AUDITING...", state=tk.DISABLED)
        self.root.update()

        try:
            res = run_audit(gen, man, lambda m: self.lbl_progress.config(text=f"● {m}"))
            self.last_out = res["output_path"]
            summary = (f"✔ Audit Complete\n\n"
                       f"Sheets: {res['sheets_audited']} | Checked: {res['invoices_checked']} | "
                       f"Mismatches: {res['mismatches']}")
            self.lbl_res.config(text=summary, fg=BRAND_BLUE)
            self.btn_open.config(state=tk.NORMAL)
            messagebox.showinfo("Success", f"Audit report created:\n{res['output_path']}")

        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            self.btn_run.config(text="   RUN AUDIT   ", state=tk.NORMAL)

    def _open_report(self):
        if hasattr(self, 'last_out'): os.startfile(self.last_out)

if __name__ == "__main__":
    root = tk.Tk()
    DSRAuditorApp(root)
    root.mainloop()
get()

        if not gen or not man:
            messagebox.showwarning("Missing Files", "Please select both files.")
            return

        # Add user-specified skip columns to the global set
        extra_skips = [norm_col(c) for c in
                       self.skip_entry.get().split(",") if c.strip()]
        SKIP_COLUMNS.update(extra_skips)

        self.btn_run.config(text="RUNNING...", state=tk.DISABLED)
        self._status("● Starting audit...", MUTED_GRAY)
        self.result_label.config(text="Audit in progress...", fg=MUTED_GRAY)
        self.btn_open.config(state=tk.DISABLED)
        self.root.update()

        try:
            result = run_audit(
                gen_path=gen,
                man_path=man,
                progress_cb=lambda m: self._status(f"● {m}", MUTED_GRAY))

            self._last_output = result["output_path"]
            checked   = result["invoices_checked"]
            mismatch  = result["mismatches"]
            unmatched = result["unmatched_gen"]
            sheets    = result["sheets_audited"]
            match_pct = round(
                (checked - mismatch) / max(checked, 1) * 100, 1)

            colour = SUCCESS_GRN if mismatch == 0 else BRAND_RED
            self._status(
                f"✔ Audit complete — {sheets} sheet(s), "
                f"{checked} invoices, {mismatch} mismatches",
                colour)

            summary = (
                f"{'✔ Clean — no mismatches!' if mismatch == 0 else f'⚠  {mismatch} mismatch(es) found'}\n\n"
                f"  Sheets audited      {sheets}\n"
                f"  Invoices matched    {checked}\n"
                f"  Mismatches          {mismatch}\n"
                f"  Not in manual DSR   {unmatched}\n"
                f"  Match rate          {match_pct}%\n\n"
                f"Red flags saved to:\n"
                f"{os.path.basename(self._last_output)}"
            )
            self.result_label.config(
                text=summary, fg=colour,
                font=("Courier", 9),
                justify="left", anchor="w",
                padx=30, pady=20)
            self.btn_open.config(state=tk.NORMAL)

        except Exception as e:
            self._status(f"✖ Error: {e}", BRAND_RED)
            messagebox.showerror("Audit Failed", str(e))

        finally:
            self.btn_run.config(text="  RUN AUDIT", state=tk.NORMAL)

    def _open_report(self):
        if self._last_output and os.path.exists(self._last_output):
            os.startfile(self._last_output)   # Windows
        else:
            messagebox.showwarning("Not Found",
                                   "Report file not found.")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = DSRAuditorApp(root)
    root.mainloop()