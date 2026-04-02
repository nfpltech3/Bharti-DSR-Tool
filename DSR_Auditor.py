import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
from datetime import datetime

class DSRAuditorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("NAGARKOT DSR AUDITOR - Data Integrity Suite")
        self.root.geometry("1100x700")
        
        # --- Nagarkot Brand System ---
        self.BRAND_RED = "#C00000"
        self.SUCCESS_GRN = "#10B981"
        self.BG_LIGHT = "#F8FAFC"
        self.TEXT_DARK = "#1E293B"
        
        self.setup_gui()
        
    def setup_gui(self):
        # Main layout
        self.root.configure(bg=self.BG_LIGHT)
        
        # 1. Header (Logo & Title)
        header = tk.Frame(self.root, bg=self.BRAND_RED, height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(header, text="NAGARKOT FORWARDERS PVT. LTD.", 
                 fg="white", bg=self.BRAND_RED, 
                 font=("Outfit", 20, "bold")).pack(pady=20)
        
        # 2. Body
        body = tk.Frame(self.root, bg=self.BG_LIGHT)
        body.pack(expand=True, fill=tk.BOTH, padx=40, pady=30)
        
        tk.Label(body, text="DSR DATA INTEGRITY AUDITOR", 
                 font=("Outfit", 18, "bold"), fg=self.TEXT_DARK, bg=self.BG_LIGHT).pack(pady=(0,30))
        
        # Selection Section
        sel_frame = tk.Frame(body, bg="white", highlightbackground="#E2E8F0", highlightthickness=1)
        sel_frame.pack(fill=tk.X, pady=10, padx=20)
        
        # File 1: Generated 
        tk.Label(sel_frame, text="1. SELECT GENERATED DSR (Source)", 
                 font=("Inter", 10, "bold"), bg="white").grid(row=0, column=0, sticky="w", padx=20, pady=(20, 5))
        self.gen_path = tk.StringVar()
        tk.Entry(sel_frame, textvariable=self.gen_path, width=80).grid(row=1, column=0, padx=20, pady=5)
        tk.Button(sel_frame, text="BROWSE...", command=lambda: self.pick_file(self.gen_path), 
                  bg=self.BRAND_RED, fg="white", width=15).grid(row=1, column=1, padx=20)
        
        # File 2: Manual
        tk.Label(sel_frame, text="2. SELECT MANUAL DSR (Reference)", 
                 font=("Inter", 10, "bold"), bg="white").grid(row=2, column=0, sticky="w", padx=20, pady=(20, 5))
        self.man_path = tk.StringVar()
        tk.Entry(sel_frame, textvariable=self.man_path, width=80).grid(row=3, column=0, padx=20, pady=5)
        tk.Button(sel_frame, text="BROWSE...", command=lambda: self.pick_file(self.man_path), 
                  bg=self.BRAND_RED, fg="white", width=15).grid(row=3, column=1, padx=20, pady=(0,20))
        
        # Audit Result Label
        self.lbl_status = tk.Label(body, text="Ready to audit.", font=("Inter", 11), fg="#64748B", bg=self.BG_LIGHT)
        self.lbl_status.pack(pady=20)
        
        # Control Buttons
        btn_frame = tk.Frame(body, bg=self.BG_LIGHT)
        btn_frame.pack(pady=20)
        
        self.btn_run = tk.Button(btn_frame, text="🔍  IDENTIFY DISCREPANCIES", 
                                 command=self.run_audit, 
                                 bg=self.BRAND_RED, fg="white", 
                                 font=("Inter", 12, "bold"), padx=30, pady=12)
        self.btn_run.pack()
        
        # Footer
        footer = tk.Frame(self.root, bg="#F1F5F9", height=40)
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        tk.Label(footer, text="© 2026 Nagarkot Forwarders Pvt Ltd. | Data Integrity Suite", 
                 fg="#94A3B8", bg="#F1F5F9", font=("Inter", 8)).pack(pady=10)

    def pick_file(self, var):
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if f: var.set(f)

    def run_audit(self):
        gen_f = self.gen_path.get()
        man_f = self.man_path.get()
        
        if not gen_f or not man_f:
            messagebox.showwarning("Incomplete", "Please select both files first.")
            return
            
        self.btn_run.config(state=tk.DISABLED, text="AUDITING...")
        self.root.update()
        
        try:
            # Load files
            xls_gen = pd.ExcelFile(gen_f)
            xls_man = pd.ExcelFile(man_f)
            
            # Setup output file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            out_file = os.path.join(os.path.dirname(gen_f), f"DSR_Audit_Report_{timestamp}.xlsx")
            
            with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                mismatch_total = 0
                
                for sheet in xls_gen.sheet_names:
                    # Match manual sheet (handle spaces)
                    man_sheet = next((s for s in xls_man.sheet_names if s.strip().lower() == sheet.strip().lower()), None)
                    if not man_sheet: continue
                    
                    df_gen = pd.read_excel(gen_f, sheet_name=sheet).fillna("")
                    df_man = pd.read_excel(man_f, sheet_name=man_sheet).fillna("")
                    
                    # Normalize columns
                    df_gen.columns = [str(c).strip().replace("\n", "").replace(" ", "") for c in df_gen.columns]
                    df_man.columns = [str(c).strip().replace("\n", "").replace(" ", "") for c in df_man.columns]
                    
                    # Target unique key: InvoiceNo
                    id_col = "InvoiceNo"
                    if id_col not in df_gen.columns or id_col not in df_man.columns:
                        continue
                        
                    # Build audit results
                    df_gen.to_excel(writer, sheet_name=sheet, index=False)
                    ws = writer.sheets[sheet]
                    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    
                    # Helper for normalization
                    def normalize_val(v):
                        v_str = str(v).strip().lower()
                        if v_str.endswith(".0"): v_str = v_str[:-2]
                        # Handle Date normalization
                        try:
                            # Try multiple date formats
                            dt = pd.to_datetime(v_str, errors='coerce')
                            if not pd.isna(dt): return dt.strftime("%Y-%m-%d")
                        except: pass
                        return v_str

                    # Match rows
                    for r_idx, gen_row in df_gen.iterrows():
                        inv_val = normalize_val(gen_row[id_col])
                        # Filter manual rows for this invoice
                        ref_row = df_man[df_man[id_col].apply(normalize_val) == inv_val]
                        
                        if not ref_row.empty:
                            ref_row = ref_row.iloc[0]
                            # Compare shared columns
                            shared_cols = set(df_gen.columns).intersection(set(df_man.columns))
                            for c_idx, col in enumerate(df_gen.columns):
                                if col in shared_cols and col != id_col:
                                    gen_cell = normalize_val(gen_row[col])
                                    man_cell = normalize_val(ref_row[col])
                                    
                                    if gen_cell != man_cell:
                                        # Highlight discrepancy (pandas row is 0-indexed, excel 1-indexed, header is row 1)
                                        cell = ws.cell(row=r_idx + 2, column=c_idx + 1)
                                        cell.fill = red_fill
                                        mismatch_total += 1
                
            self.lbl_status.config(text=f"✔ Success! {mismatch_total} discrepancies identified.", fg=self.SUCCESS_GRN)
            messagebox.showinfo("Audit Complete", f"Found {mismatch_total} discrepancies.\n\nCreated Audit Report:\n{out_file}")
            
        except Exception as e:
            messagebox.showerror("Audit Error", f"Failed to perform audit:\n{e}")
            self.lbl_status.config(text="✖ Audit failed.", fg=self.BRAND_RED)
            
        finally:
            self.btn_run.config(state=tk.NORMAL, text="🔍  IDENTIFY DISCREPANCIES")

if __name__ == "__main__":
    root = tk.Tk()
    DSRAuditorApp(root)
    root.mainloop()
